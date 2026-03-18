"""Generate an Excel workbook from extracted invoice data."""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")

FIELD_LABELS = {
    "invoice_number": "Invoice Number",
    "invoice_date": "Invoice Date",
    "supplier_name": "Supplier Name",
    "supplier_address": "Supplier Address",
    "client_name": "Client Name",
    "client_address": "Client Address",
    "supplier_tax_id": "Supplier Tax ID",
    "client_tax_id": "Client Tax ID",
    "total_excl_vat": "Total (excl. VAT)",
    "total_incl_vat": "Total (incl. VAT)",
    "currency": "Currency",
    "vat_rate": "VAT Rate",
    "vat_amount": "VAT Amount",
}

LINE_ITEM_COLUMNS = [
    ("description", "Description"),
    ("quantity", "Quantity"),
    ("unit_price", "Unit Price"),
    ("unit", "Unit"),
    ("total", "Total"),
]


def create_excel(data: dict) -> bytes:
    """Create an Excel workbook with two sheets: fields and line items.

    Returns:
        Raw bytes of the .xlsx file, ready to send.
    """
    wb = Workbook()

    # --- Sheet 1: Invoice Summary ---
    ws_fields = wb.active
    ws_fields.title = "Invoice Summary"
    ws_fields.column_dimensions["A"].width = 22
    ws_fields.column_dimensions["B"].width = 50

    # Header
    ws_fields.append(["Field", "Value"])
    for cell in ws_fields[1]:
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left")

    fields = data.get("fields", {})
    for key, label in FIELD_LABELS.items():
        value = fields.get(key)
        ws_fields.append([label, _format_value(value)])
        ws_fields.cell(row=ws_fields.max_row, column=2).alignment = Alignment(
            horizontal="left"
        )

    # Additional fields (from format-flexible extraction)
    for key, val in data.get("additional_fields", {}).items():
        label = key.replace("_", " ").title()
        ws_fields.append([label, _format_value(val)])
        ws_fields.cell(row=ws_fields.max_row, column=2).alignment = Alignment(
            horizontal="left"
        )

    # --- Sheet 2: Line Items ---
    ws_items = wb.create_sheet("Line Items")
    col_keys = [c[0] for c in LINE_ITEM_COLUMNS]
    col_headers = [c[1] for c in LINE_ITEM_COLUMNS]

    # Detect extra line item keys for format flexibility
    known_keys = set(col_keys)
    for item in data.get("line_items", []):
        for key in item:
            if key not in known_keys:
                known_keys.add(key)
                col_keys.append(key)
                col_headers.append(key.replace("_", " ").title())

    ws_items.append(col_headers)
    for cell in ws_items[1]:
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left")

    # Column widths
    base_widths = [40, 12, 14, 10, 14]
    widths = base_widths + [14] * (len(col_keys) - len(base_widths))
    for i, w in enumerate(widths, 1):
        ws_items.column_dimensions[get_column_letter(i)].width = w

    for item in data.get("line_items", []):
        row = [_format_value(item.get(k)) for k in col_keys]
        ws_items.append(row)
        for cell in ws_items[ws_items.max_row]:
            cell.alignment = Alignment(horizontal="left")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _format_value(value):
    if value is None:
        return "N/A"
    return value
