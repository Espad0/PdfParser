import io

from openpyxl import load_workbook

from excel_export import create_excel


def _make_data():
    return {
        "fields": {
            "invoice_number": "INV-001",
            "invoice_date": "2024-01-15",
            "supplier_name": "Acme Corp",
            "supplier_address": "123 Main St",
            "client_name": "Client Inc",
            "client_address": "456 Oak Ave",
            "supplier_tax_id": "DE123456789",
            "client_tax_id": "FR987654321",
            "total_excl_vat": 1000.0,
            "total_incl_vat": 1200.0,
            "currency": "EUR",
            "vat_rate": "20%",
            "vat_amount": 200.0,
        },
        "additional_fields": {},
        "line_items": [
            {"description": "Widget", "quantity": 10, "unit_price": 100.0, "unit": "pc", "total": 1000.0},
        ],
    }


def _load(data):
    return load_workbook(io.BytesIO(create_excel(data)))


class TestCreateExcel:
    def test_returns_valid_xlsx(self):
        result = create_excel(_make_data())
        assert isinstance(result, bytes)
        wb = load_workbook(io.BytesIO(result))
        assert len(wb.sheetnames) == 2

    def test_sheet_names(self):
        wb = _load(_make_data())
        assert wb.sheetnames == ["Invoice Summary", "Line Items"]

    def test_field_values_in_summary(self):
        wb = _load(_make_data())
        ws = wb["Invoice Summary"]
        values = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
                  for r in range(2, ws.max_row + 1)}
        assert values["Invoice Number"] == "INV-001"
        assert values["Total (excl. VAT)"] == 1000.0

    def test_line_items_in_sheet(self):
        wb = _load(_make_data())
        ws = wb["Line Items"]
        assert ws.cell(row=1, column=1).value == "Description"
        assert ws.cell(row=2, column=1).value == "Widget"
        assert ws.cell(row=2, column=2).value == 10

    def test_empty_line_items(self):
        data = _make_data()
        data["line_items"] = []
        wb = _load(data)
        ws = wb["Line Items"]
        assert ws.max_row == 1  # header only

    def test_none_values_shown_as_na(self):
        data = _make_data()
        data["fields"]["invoice_number"] = None
        wb = _load(data)
        ws = wb["Invoice Summary"]
        values = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
                  for r in range(2, ws.max_row + 1)}
        assert values["Invoice Number"] == "N/A"

    def test_additional_fields_in_summary(self):
        data = _make_data()
        data["additional_fields"] = {"payment_terms": "Net 30"}
        wb = _load(data)
        ws = wb["Invoice Summary"]
        values = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
                  for r in range(2, ws.max_row + 1)}
        assert values["Payment Terms"] == "Net 30"

    def test_extra_line_item_keys_as_columns(self):
        data = _make_data()
        data["line_items"] = [
            {"description": "X", "quantity": 1, "unit_price": 10.0, "unit": "pc", "total": 10.0, "item_code": "A1"},
        ]
        wb = _load(data)
        ws = wb["Line Items"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Item Code" in headers
